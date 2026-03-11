"""
Chili Prices API
================

Endpoints for chili price tracking and history.
Admin-only write access, public read access.
"""

from datetime import datetime, timedelta
from typing import List, Optional
from fastapi import APIRouter, Depends, HTTPException, Query, status
from bson import ObjectId

from ..core.database import MongoDB
from ..core.security import require_admin
from ..schemas.price import (
    PriceEntryCreate,
    PriceEntryUpdate,
    PriceEntryResponse,
    PriceAnalytics,
    CurrentPrices,
    PriceComparisonResponse,
    PricePredictionResponse,
    PriceTrend,
    ChiliType,
    MarketLocation,
)
from ..ml.price_predictor import price_predictor

router = APIRouter(prefix="/prices", tags=["Chili Prices"])


def price_doc_to_response(doc: dict) -> PriceEntryResponse:
    """Convert MongoDB document to response schema."""
    return PriceEntryResponse(
        id=str(doc["_id"]),
        chili_type=doc["chili_type"],
        price=doc["price"],
        unit=doc["unit"],
        location=doc["location"],
        market_name=doc.get("market_name"),
        notes=doc.get("notes"),
        recorded_date=doc["recorded_date"],
        created_by=doc["created_by"],
        created_at=doc["created_at"],
        updated_at=doc.get("updated_at"),
    )


@router.post("", response_model=PriceEntryResponse, status_code=status.HTTP_201_CREATED)
async def create_price_entry(
    entry: PriceEntryCreate,
    current_user: dict = Depends(require_admin),
):
    """
    Create a new price entry. Admin only.
    """
    db = MongoDB.database
    
    now = datetime.utcnow()
    doc = {
        "chili_type": entry.chili_type.value,
        "price": entry.price,
        "unit": entry.unit.value,
        "location": entry.location.value,
        "market_name": entry.market_name,
        "notes": entry.notes,
        "recorded_date": entry.recorded_date or now,
        "created_by": current_user["user_id"],
        "created_at": now,
    }
    
    result = await db.chili_prices.insert_one(doc)
    doc["_id"] = result.inserted_id
    
    return price_doc_to_response(doc)


@router.get("", response_model=List[PriceEntryResponse])
async def get_price_history(
    chili_type: Optional[ChiliType] = None,
    location: Optional[MarketLocation] = None,
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
    limit: int = Query(default=100, le=500),
):
    """
    Get price history with optional filters.
    """
    db = MongoDB.database
    
    query = {}
    if chili_type:
        query["chili_type"] = chili_type.value
    if location:
        query["location"] = location.value
    if start_date or end_date:
        query["recorded_date"] = {}
        if start_date:
            query["recorded_date"]["$gte"] = start_date
        if end_date:
            query["recorded_date"]["$lte"] = end_date
    
    cursor = db.chili_prices.find(query).sort("recorded_date", -1).limit(limit)
    docs = await cursor.to_list(length=limit)
    
    return [price_doc_to_response(doc) for doc in docs]


@router.get("/current", response_model=CurrentPrices)
async def get_current_prices(
    location: MarketLocation = MarketLocation.NATIONAL_AVERAGE,
):
    """
    Get the most recent prices for all chili types.
    """
    db = MongoDB.database
    
    prices = {}
    last_updated = None
    
    for chili_type in ChiliType:
        # Get the most recent price for this type and location
        doc = await db.chili_prices.find_one(
            {"chili_type": chili_type.value, "location": location.value},
            sort=[("recorded_date", -1)]
        )
        if doc:
            prices[chili_type.value] = doc["price"]
            if not last_updated or doc["recorded_date"] > last_updated:
                last_updated = doc["recorded_date"]
    
    return CurrentPrices(
        siling_haba=prices.get("siling_haba"),
        siling_labuyo=prices.get("siling_labuyo"),
        siling_demonyo=prices.get("siling_demonyo"),
        location=location,
        last_updated=last_updated or datetime.utcnow(),
    )


@router.get("/market-overview")
async def get_market_overview():
    """
    Full market overview: all 3 varieties with price trend, stats, and latest prices.
    Designed for the frontend Market Prices dashboard.
    """
    db = MongoDB.database

    overview = {}

    for ctype in ["siling_labuyo", "siling_haba", "siling_demonyo"]:
        # All prices sorted by date
        cursor = db.chili_prices.find(
            {"chili_type": ctype}
        ).sort("recorded_date", 1)
        docs = await cursor.to_list(length=500)

        if not docs:
            overview[ctype] = {
                "current_price": 0, "current_low": 0, "current_high": 0,
                "avg_price": 0, "min_price": 0, "max_price": 0,
                "price_change_pct": 0, "trend": [], "data_points": 0,
            }
            continue

        # Build trend array with low/high
        trend = []
        for doc in docs:
            trend.append({
                "date": doc["recorded_date"].strftime("%Y-%m-%d"),
                "price": doc["price"],
                "low": doc.get("price_low", doc["price"]),
                "high": doc.get("price_high", doc["price"]),
            })

        prices = [d["price"] for d in docs]
        latest = docs[-1]
        earliest = docs[0]
        change_pct = 0
        if earliest["price"] > 0:
            change_pct = round((latest["price"] - earliest["price"]) / earliest["price"] * 100, 1)

        overview[ctype] = {
            "current_price": latest["price"],
            "current_low": latest.get("price_low", latest["price"]),
            "current_high": latest.get("price_high", latest["price"]),
            "avg_price": round(sum(prices) / len(prices), 2),
            "min_price": min(prices),
            "max_price": max(prices),
            "price_change_pct": change_pct,
            "data_points": len(docs),
            "date_range": {
                "start": docs[0]["recorded_date"].strftime("%Y-%m-%d"),
                "end": docs[-1]["recorded_date"].strftime("%Y-%m-%d"),
            },
            "trend": trend,
        }

    return overview


@router.get("/analytics/{chili_type}", response_model=PriceAnalytics)
async def get_price_analytics(
    chili_type: ChiliType,
    location: MarketLocation = MarketLocation.NATIONAL_AVERAGE,
):
    """
    Get price analytics and trends for a specific chili type.
    """
    db = MongoDB.database
    
    now = datetime.utcnow()
    date_30d = now - timedelta(days=30)
    date_7d = now - timedelta(days=7)
    
    # Get current price
    current_doc = await db.chili_prices.find_one(
        {"chili_type": chili_type.value, "location": location.value},
        sort=[("recorded_date", -1)]
    )
    current_price = current_doc["price"] if current_doc else 0
    
    # Get 30-day stats
    pipeline_30d = [
        {
            "$match": {
                "chili_type": chili_type.value,
                "location": location.value,
                "recorded_date": {"$gte": date_30d}
            }
        },
        {
            "$group": {
                "_id": None,
                "avg_price": {"$avg": "$price"},
                "min_price": {"$min": "$price"},
                "max_price": {"$max": "$price"},
            }
        }
    ]
    
    stats_30d = await db.chili_prices.aggregate(pipeline_30d).to_list(1)
    stats = stats_30d[0] if stats_30d else {"avg_price": 0, "min_price": 0, "max_price": 0}
    
    # Get price 7 days ago for change calculation
    doc_7d = await db.chili_prices.find_one(
        {
            "chili_type": chili_type.value,
            "location": location.value,
            "recorded_date": {"$lte": date_7d}
        },
        sort=[("recorded_date", -1)]
    )
    price_7d_ago = doc_7d["price"] if doc_7d else current_price
    
    # Get price 30 days ago
    doc_30d = await db.chili_prices.find_one(
        {
            "chili_type": chili_type.value,
            "location": location.value,
            "recorded_date": {"$lte": date_30d}
        },
        sort=[("recorded_date", -1)]
    )
    price_30d_ago = doc_30d["price"] if doc_30d else current_price
    
    # Calculate changes
    change_7d = ((current_price - price_7d_ago) / price_7d_ago * 100) if price_7d_ago else 0
    change_30d = ((current_price - price_30d_ago) / price_30d_ago * 100) if price_30d_ago else 0
    
    # Get daily trend for last 30 days
    pipeline_trend = [
        {
            "$match": {
                "chili_type": chili_type.value,
                "location": location.value,
                "recorded_date": {"$gte": date_30d}
            }
        },
        {
            "$group": {
                "_id": {
                    "$dateToString": {"format": "%Y-%m-%d", "date": "$recorded_date"}
                },
                "average_price": {"$avg": "$price"},
                "min_price": {"$min": "$price"},
                "max_price": {"$max": "$price"},
                "sample_count": {"$sum": 1}
            }
        },
        {"$sort": {"_id": 1}}
    ]
    
    trend_docs = await db.chili_prices.aggregate(pipeline_trend).to_list(30)
    trend = [
        PriceTrend(
            date=datetime.strptime(doc["_id"], "%Y-%m-%d"),
            average_price=doc["average_price"],
            min_price=doc["min_price"],
            max_price=doc["max_price"],
            sample_count=doc["sample_count"]
        )
        for doc in trend_docs
    ]
    
    return PriceAnalytics(
        chili_type=chili_type,
        location=location,
        current_price=current_price,
        price_change_7d=round(change_7d, 2),
        price_change_30d=round(change_30d, 2),
        avg_price_30d=round(stats["avg_price"], 2),
        min_price_30d=round(stats["min_price"], 2),
        max_price_30d=round(stats["max_price"], 2),
        trend=trend,
    )


@router.get("/compare/{chili_type}", response_model=PriceComparisonResponse)
async def compare_prices_by_location(
    chili_type: ChiliType,
):
    """
    Compare prices across different locations for a chili type.
    """
    db = MongoDB.database
    
    # Get latest price for each location
    pipeline = [
        {"$match": {"chili_type": chili_type.value}},
        {"$sort": {"recorded_date": -1}},
        {
            "$group": {
                "_id": "$location",
                "price": {"$first": "$price"}
            }
        }
    ]
    
    docs = await db.chili_prices.aggregate(pipeline).to_list(20)
    
    prices_by_location = {doc["_id"]: doc["price"] for doc in docs}
    
    if not prices_by_location:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="No price data found for this chili type"
        )
    
    prices = list(prices_by_location.values())
    national_avg = sum(prices) / len(prices)
    
    cheapest = min(prices_by_location, key=prices_by_location.get)
    most_expensive = max(prices_by_location, key=prices_by_location.get)
    
    return PriceComparisonResponse(
        chili_type=chili_type,
        prices_by_location=prices_by_location,
        national_average=round(national_avg, 2),
        cheapest_location=cheapest,
        most_expensive_location=most_expensive,
    )


@router.put("/{price_id}", response_model=PriceEntryResponse)
async def update_price_entry(
    price_id: str,
    update: PriceEntryUpdate,
    current_user: dict = Depends(require_admin),
):
    """
    Update a price entry. Admin only.
    """
    db = MongoDB.database
    
    try:
        oid = ObjectId(price_id)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid price ID")
    
    update_data = {k: v for k, v in update.model_dump().items() if v is not None}
    if not update_data:
        raise HTTPException(status_code=400, detail="No update data provided")
    
    update_data["updated_at"] = datetime.utcnow()
    
    result = await db.chili_prices.find_one_and_update(
        {"_id": oid},
        {"$set": update_data},
        return_document=True
    )
    
    if not result:
        raise HTTPException(status_code=404, detail="Price entry not found")
    
    return price_doc_to_response(result)


@router.delete("/{price_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_price_entry(
    price_id: str,
    current_user: dict = Depends(require_admin),
):
    """
    Delete a price entry. Admin only.
    """
    db = MongoDB.database
    
    try:
        oid = ObjectId(price_id)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid price ID")
    
    result = await db.chili_prices.delete_one({"_id": oid})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Price entry not found")


@router.get("/predict/model-info")
async def get_prediction_model_info():
    """
    Get information about the trained price prediction model.
    """
    if not price_predictor.is_loaded:
        price_predictor.load()
    return price_predictor.get_model_info()


@router.get("/predict/{chili_type}", response_model=PricePredictionResponse)
async def predict_prices(
    chili_type: ChiliType,
    days: int = Query(default=7, ge=1, le=90, description="Days to forecast (1-90)"),
):
    """
    Predict future chili prices using the trained Random Forest model.
    
    Returns price forecasts for the next N days (1-90) for the specified chili type.
    """
    if not price_predictor.is_loaded:
        if not price_predictor.load():
            raise HTTPException(
                status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
                detail="Price prediction model is not available. Please train the model first.",
            )
    
    try:
        result = price_predictor.predict(
            chili_type=chili_type.value,
            days_ahead=days,
        )
        return result
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Prediction error: {str(e)}",
        )


@router.post("/seed", status_code=status.HTTP_201_CREATED)
async def seed_sample_prices(
    current_user: dict = Depends(require_admin),
):
    """
    Seed database with real price data from Excel files. Admin only.

    Reads directly from:
    - assets/Siling Labuyo Price(3).xlsx — 439 records (Jan 2024 – Dec 2025), LOW/HIGH
    - assets/Siling Haba.xlsx — 141 records (Sep 2025 – Feb 2026), PREVAILING
    """
    import os
    import openpyxl

    db = MongoDB.database

    # Clear existing price data to avoid duplicates
    await db.chili_prices.delete_many({})

    now = datetime.utcnow()
    user_id = current_user["user_id"]
    sample_data: list[dict] = []

    # Resolve the assets directory (relative to backend/)
    assets_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))), "..", "assets")
    assets_dir = os.path.normpath(assets_dir)

    # ----------------------------------------------------------------
    # REAL DATA — Siling Labuyo from Excel
    # 439 records (Jan 2024 – Dec 2025) with LOW, HIGH columns
    # ----------------------------------------------------------------
    labuyo_path = os.path.join(assets_dir, "Siling Labuyo Price(3).xlsx")
    labuyo_count = 0

    if os.path.exists(labuyo_path):
        wb = openpyxl.load_workbook(labuyo_path, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        for row in rows[2:]:  # Skip 2-row header
            date_val = row[0]
            if date_val is None:
                continue
            # Skip year-only integers (mid-file sub-headers like 2024, 2025)
            if isinstance(date_val, (int, float)) and not isinstance(date_val, datetime):
                continue
            if not isinstance(date_val, datetime):
                try:
                    from dateutil.parser import parse as date_parse
                    date_val = date_parse(str(date_val))
                except Exception:
                    continue

            low_val = row[1]
            high_val = row[2]
            prevailing_val = row[3] if len(row) > 3 else None
            average_val = row[4] if len(row) > 4 else None

            # Safely convert to float (skip rows with non-numeric values like headers)
            def _safe_float(val):
                if val is None:
                    return None
                try:
                    return float(val)
                except (ValueError, TypeError):
                    return None

            # Determine price values
            low = _safe_float(low_val)
            high = _safe_float(high_val)
            prevailing = _safe_float(prevailing_val)
            average = _safe_float(average_val)

            # Compute best average price
            if average and average > 0:
                avg_price = average
            elif prevailing and prevailing > 0:
                avg_price = prevailing
            elif low and high and low > 0 and high > 0:
                avg_price = (low + high) / 2
            elif low and low > 0:
                avg_price = low
            elif high and high > 0:
                avg_price = high
            else:
                continue  # No valid price

            price_low = low if (low and low > 0) else round(avg_price * 0.8)
            price_high = high if (high and high > 0) else round(avg_price * 1.2)

            sample_data.append({
                "chili_type": "siling_labuyo",
                "price": round(avg_price, 2),
                "unit": "per_kilo",
                "location": "metro_manila",
                "market_name": "Metro Manila Markets",
                "notes": f"Low: ₱{price_low:.0f} | High: ₱{price_high:.0f}",
                "recorded_date": date_val,
                "created_by": user_id,
                "created_at": now,
                "price_low": price_low,
                "price_high": price_high,
            })
            labuyo_count += 1

    # ----------------------------------------------------------------
    # REAL DATA — Siling Haba from Excel
    # 141 records (Sep 2025 – Feb 2026) with PREVAILING prices
    # ----------------------------------------------------------------
    haba_path = os.path.join(assets_dir, "Siling Haba.xlsx")
    haba_count = 0

    if os.path.exists(haba_path):
        wb = openpyxl.load_workbook(haba_path, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        for row in rows[2:]:  # Skip 2-row header
            date_val = row[0]
            if date_val is None:
                continue
            # Skip year-only integers (mid-file sub-headers)
            if isinstance(date_val, (int, float)) and not isinstance(date_val, datetime):
                continue
            if not isinstance(date_val, datetime):
                try:
                    from dateutil.parser import parse as date_parse
                    date_val = date_parse(str(date_val))
                except Exception:
                    continue

            low_val = row[1]
            high_val = row[2]
            prevailing_val = row[3] if len(row) > 3 else None
            average_val = row[4] if len(row) > 4 else None

            # Safely convert to float (skip rows with non-numeric values)
            def safe_float(val):
                if val is None:
                    return None
                try:
                    return float(val)
                except (ValueError, TypeError):
                    return None

            # Determine price values
            low = safe_float(low_val)
            high = safe_float(high_val)
            prevailing = safe_float(prevailing_val)
            average = safe_float(average_val)

            # Compute best average price
            if average and average > 0:
                avg_price = average
            elif prevailing and prevailing > 0:
                avg_price = prevailing
            elif low and high and low > 0 and high > 0:
                avg_price = (low + high) / 2
            elif low and low > 0:
                avg_price = low
            elif high and high > 0:
                avg_price = high
            else:
                continue

            price_low = low if (low and low > 0) else round(avg_price * 0.85)
            price_high = high if (high and high > 0) else round(avg_price * 1.15)

            sample_data.append({
                "chili_type": "siling_haba",
                "price": round(avg_price, 2),
                "unit": "per_kilo",
                "location": "metro_manila",
                "market_name": "Metro Manila Markets",
                "notes": f"Prevailing: ₱{avg_price:.2f}",
                "recorded_date": date_val,
                "created_by": user_id,
                "created_at": now,
                "price_low": price_low,
                "price_high": price_high,
            })
            haba_count += 1

    # ----------------------------------------------------------------
    # ESTIMATED — Siling Demonyo (derived from Labuyo data)
    # No real Excel source yet. Estimated at ~1.45× Labuyo prices.
    # ----------------------------------------------------------------
    labuyo_entries = [d for d in sample_data if d["chili_type"] == "siling_labuyo"]
    demonyo_count = 0
    for entry in labuyo_entries:
        multiplier = 1.45
        low_d = round(entry["price_low"] * multiplier)
        high_d = round(entry["price_high"] * multiplier)
        avg_d = round((low_d + high_d) / 2, 2)
        sample_data.append({
            "chili_type": "siling_demonyo",
            "price": avg_d,
            "unit": "per_kilo",
            "location": "metro_manila",
            "market_name": "Metro Manila Markets (estimated)",
            "notes": f"Low: ₱{low_d} | High: ₱{high_d} (estimated from Labuyo ×1.45)",
            "recorded_date": entry["recorded_date"],
            "created_by": user_id,
            "created_at": now,
            "price_low": low_d,
            "price_high": high_d,
        })
        demonyo_count += 1

    if not sample_data:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"No Excel data found. Looked in: {assets_dir}",
        )

    result = await db.chili_prices.insert_many(sample_data)

    # Compute date ranges per type
    labuyo_dates = [d["recorded_date"] for d in sample_data if d["chili_type"] == "siling_labuyo"]
    haba_dates = [d["recorded_date"] for d in sample_data if d["chili_type"] == "siling_haba"]
    demonyo_dates = [d["recorded_date"] for d in sample_data if d["chili_type"] == "siling_demonyo"]

    return {
        "message": f"Seeded {len(result.inserted_ids)} price entries from Excel files",
        "breakdown": {
            "siling_labuyo": labuyo_count,
            "siling_haba": haba_count,
            "siling_demonyo": demonyo_count,
        },
        "date_ranges": {
            "siling_labuyo": f"{min(labuyo_dates).strftime('%Y-%m-%d')} to {max(labuyo_dates).strftime('%Y-%m-%d')}" if labuyo_dates else "N/A",
            "siling_haba": f"{min(haba_dates).strftime('%Y-%m-%d')} to {max(haba_dates).strftime('%Y-%m-%d')}" if haba_dates else "N/A",
            "siling_demonyo": f"{min(demonyo_dates).strftime('%Y-%m-%d')} to {max(demonyo_dates).strftime('%Y-%m-%d')}" if demonyo_dates else "N/A",
        },
        "note": "Labuyo & Haba from real Excel sources; Demonyo estimated from Labuyo (×1.45)",
    }
